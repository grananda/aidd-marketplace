#!/usr/bin/env python3
"""aiba-test-plan · gen_test_plan_xlsx.py — Plan de Pruebas de una HU.

Renderiza un manifiesto JSON al `.xlsx` del plan. Hay un unico generador a
proposito, por la misma razon que en el DF: el valor del skill es que todos los
planes de un proyecto salgan iguales, y eso no se sostiene si cada invocacion
arma el libro a su manera.

Seis hojas: Hoja de Control, Especificaciones (el inventario), Parametros,
Ejecucion 1, Qmetry y Resumen.

**Sin macros.** La plantilla de referencia es un `.xlsm` cuyos botones generan
los codigos y montan la hoja de ejecucion; aqui eso sale hecho al construir el
fichero. Lo que si se conserva son sus reglas: el correlativo reinicia por
prefijo, los tramos vacios se colapsan, y dos casos que caen en el mismo codigo
se avisan en vez de numerarse en silencio.

Uso:
    python3 gen_test_plan_xlsx.py --manifest plan.json --output "docs/pruebas/PP - HU-06.xlsx"
    python3 gen_test_plan_xlsx.py --schema

Solo depende de `openpyxl`, que instala sobre la marcha si falta (misma
estrategia que `gen_df_docx.py`: la invocacion de python es la puerta de
permisos y el pip hereda esa aprobacion). Se desactiva con `--no-install` o la
variable de entorno `AIBA_PP_NO_INSTALL`.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import date
from pathlib import Path

# El modulo de marca vive a nivel de plugin: skills/<skill>/scripts/ -> parents[3].
sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "scripts"))

# --- Dependencia -------------------------------------------------------------

def _ensure_openpyxl(allow_install: bool) -> None:
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass

    if not allow_install or os.environ.get("AIBA_PP_NO_INSTALL"):
        sys.stderr.write(
            "Falta 'openpyxl' y la instalacion automatica esta desactivada.\n"
            "Instalalo con:  pip install openpyxl\n")
        sys.exit(2)

    sys.stderr.write("Aviso: 'openpyxl' no esta instalado; instalandolo automaticamente...\n")
    for cmd in ([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
                [sys.executable, "-m", "pip", "install", "--quiet", "--user", "openpyxl"]):
        try:
            subprocess.check_call(cmd)
        except Exception:  # noqa: BLE001 - se prueba la siguiente estrategia
            continue
        try:
            import openpyxl  # noqa: F401
            sys.stderr.write("OK: 'openpyxl' instalado correctamente.\n")
            return
        except ImportError:
            continue
    sys.stderr.write("No se pudo instalar 'openpyxl'. Instalalo a mano y reintenta.\n")
    sys.exit(2)


SCHEMA = """\
Esquema del manifiesto (JSON). Lo que falte se omite o sale vacio.

{
  "proyecto":   "FUS - Frontal Unico Suplementos",
  "hu_id":      "HU-06",
  "titulo":     "Tomador: visualizacion y modificacion",
  "referencia": "STRY0113597",          # opcional: id de la HU en Jira/Qmetry
  "version":    "1.0",
  "autor":      "Nombre Apellido",
  "fecha":      "2026-08-29",           # opcional; por defecto, hoy
  "entorno":    "DEV",                  # entorno previsto de ejecucion
  "resumen":    "Una linea sobre el alcance de este plan",

  "control_versiones": [{"version":"1.0","descripcion":"Version inicial",
                         "autor":"...","fecha":"2026-08-29"}],

  "casos_uso": [                        # agrupan los casos de prueba
    {"id":"CU01", "nombre":"Acceso a pantalla Tomador",
     "descripcion":"...", "requisitos":["RF-12"]}
  ],

  "casos": [
    {"caso_uso":"CU01",                 # debe existir en casos_uso
     "nombre":"Acceso y visualizacion de Tomador",
     "criticidad":"Alta",               # Alta | Media | Baja
     "descripcion":"Que verifica y por que",
     "precondiciones":"Estado previo necesario",
     "pasos":["1. ...","2. ..."],       # texto o lista
     "resultado_esperado":"Que debe ocurrir",
     "requisitos":["RF-12","NFR-03"],
     "ejecucion":"manual",              # manual | automatizable
     "change":"portal-catalogo",        # change_hint del roadmap, si existe
     "duracion":"5 min"}
  ],

  "branding": {                         # opcional; sin el, libro neutro
    "color_principal":"1F3864", "color_secundario":"2E74B5",
    "logo":"ruta/al/logo.png", "texto_cabecera":"...", "texto_pie":"..."
  }
}
"""

NIVEL = "Pruebas de Sistema-Funcionales"
NIVEL_COD = ("PS", "FU")

CRITICIDADES = ["Alta", "Media", "Baja"]
ESTADOS = ["Pendiente", "En ejecución", "Pasado", "Fallado", "Bloqueado"]
EJECUCIONES = ["manual", "automatizable"]
RESULTADOS = ["No ejecutado", "Pasado", "Fallado", "Bloqueado"]

# Columnas del inventario, en el orden y con los nombres de la plantilla de
# referencia. La columna I es `Estado` y no `Version`: en la plantilla existe
# para que la macro distinga versiones de duplicados, y sin macro no aporta.
COLUMNAS = [
    ("Nivel de Prueba", 24), ("<Nivel de Agrupación 1>", 26),
    ("<Nivel de Agrupación 2>", 30), ("<Nivel de Agrupación 3>", 18),
    ("Código", 18), ("Nombre del caso de prueba", 38), ("Criticidad", 11),
    ("Estado", 13), ("Descripción", 52), ("Precondiciones", 38), ("Pasos", 52),
    ("Resultados Esperados", 52), ("Requisitos relacionados", 22),
    ("Keywords", 16), ("ID Externo", 20), ("Duración", 11),
    ("Selección carga", 14),
]
COL_INI = 2          # el inventario arranca en B, como la plantilla
FILA_CAB = 7         # cabeceras
FILA_DATOS = 8       # primer caso


# --- Codigos -----------------------------------------------------------------

def _tramo(valor: str | None) -> str:
    """Un tramo del codigo. `NA` y sus variantes valen como vacio."""
    v = (valor or "").strip()
    if v.upper() in ("", "NA", "N/A", "NO APLICA"):
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", v).upper()


def generar_codigos(casos: list[dict]) -> list[str]:
    """`PS.FU.<CU>.<nn>`, con el correlativo reiniciado por prefijo.

    Reglas heredadas de la macro `A1_GenerateCodes` de la plantilla: los tramos
    vacios se colapsan (no quedan puntos dobles) y el correlativo cuenta dentro
    de cada prefijo, no sobre el total.
    """
    contador: dict[str, int] = {}
    codigos: list[str] = []
    for c in casos:
        tramos = [NIVEL_COD[0], NIVEL_COD[1], _tramo(c.get("caso_uso"))]
        prefijo = ".".join(t for t in tramos if t)
        contador[prefijo] = contador.get(prefijo, 0) + 1
        codigos.append(f"{prefijo}.{contador[prefijo]:02d}")
    return codigos


def detectar_colisiones(casos: list[dict], codigos: list[str]) -> list[str]:
    """Casos que comparten caso de uso y nombre: casi siempre un duplicado real.

    La plantilla los pintaba en rojo al pulsar el boton. Aqui se avisa al
    generar, que es cuando todavia se puede arreglar en el origen.
    """
    vistos: dict[tuple[str, str], str] = {}
    avisos: list[str] = []
    for c, cod in zip(casos, codigos):
        clave = (_tramo(c.get("caso_uso")), (c.get("nombre") or "").strip().lower())
        if clave in vistos:
            avisos.append(
                f"{cod} y {vistos[clave]} tienen el mismo caso de uso y el mismo "
                f"nombre ('{c.get('nombre')}'): revisa si son el mismo caso duplicado")
        else:
            vistos[clave] = cod
    return avisos


# --- Utilidades de contenido -------------------------------------------------

# Excel rechaza los caracteres de control (salvo tabulador y salto de linea) y
# aborta la escritura con IllegalCharacterError. Llegan solos: los documentos de
# origen se escriben copiando y pegando de Word o de un PDF, que arrastran
# tabuladores verticales y saltos de pagina invisibles. Perder el fichero entero
# por un caracter que nadie ve, y ademas con una traza de Python, no es opcion.
CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
LIMITE_CELDA = 32767  # el maximo de una celda de Excel

_TRUNCADOS: list[str] = []


def sanear(obj):
    """Quita los caracteres de control de **todo** el manifiesto, de una vez.

    En la entrada y no en cada escritura: hay una docena de sitios que llevan un
    valor del manifiesto a una celda, y basta que uno se olvide de limpiarlo para
    que openpyxl aborte con una traza. Saneando aqui, olvidarse deja de ser
    posible.
    """
    if isinstance(obj, str):
        return CONTROL.sub("", obj)
    if isinstance(obj, dict):
        return {k: sanear(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [sanear(v) for v in obj]
    return obj


def texto(valor) -> str:
    """Texto suelto o lista -> una cadena con saltos de linea, apta para Excel."""
    if valor is None:
        return ""
    if isinstance(valor, (list, tuple)):
        bruto = "\n".join(texto(v) for v in valor if v is not None and str(v).strip())
    else:
        bruto = str(valor).strip()
    limpio = bruto
    if len(limpio) > LIMITE_CELDA:
        # openpyxl corta por su cuenta y en silencio. Cortar aqui es lo mismo,
        # pero deja constancia: un caso al que le falta la mitad de los pasos
        # parece completo hasta que alguien intenta ejecutarlo.
        _TRUNCADOS.append(limpio[:60] + "...")
        limpio = limpio[: LIMITE_CELDA - 20] + "\n[TEXTO TRUNCADO]"
    return limpio


def lista(valor) -> str:
    """Lista de ids -> `RF-12, NFR-03`."""
    if valor is None:
        return ""
    if isinstance(valor, (list, tuple)):
        bruto = ", ".join(str(v).strip() for v in valor if str(v).strip())
    else:
        bruto = str(valor).strip()
    return CONTROL.sub("", bruto)


# --- Hojas -------------------------------------------------------------------

def _validacion(ws, valores: list[str], rango: str) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(type="list", formula1='"' + ",".join(valores) + '"',
                        allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add(rango)


def hoja_control(wb, m: dict, est: dict, br: dict, hoy: str, marca) -> None:
    """Portada y registro de modificaciones. Es lo que se ve al abrir."""
    ws = wb.active
    ws.title = "Hoja de Control"
    ws.sheet_view.showGridLines = False
    for col, ancho in (("A", 3), ("B", 26), ("C", 40), ("D", 22), ("E", 22),
                       ("F", 20), ("G", 20)):
        ws.column_dimensions[col].width = ancho

    logo = marca.logo_excel(ws, br, "B2") if br.get("logo") else False
    fila = 6 if logo else 2

    ws.cell(fila, 2, "PLAN DE PRUEBAS").font = est["titulo_fuente"]
    fila += 2
    datos = [
        ("Proyecto", texto(m.get("proyecto"))),
        ("Documento", f"Plan de Pruebas · {m.get('hu_id','')} · {m.get('titulo','')}".strip(" ·")),
        ("Historia de usuario", texto(m.get("hu_id"))),
        ("Referencia externa", texto(m.get("referencia"))),
        ("Entorno previsto", texto(m.get("entorno"))),
        ("Versión", str(m.get("version", "1.0"))),
        ("Autor", texto(m.get("autor"))),
        ("Fecha", hoy),
        ("Resumen", texto(m.get("resumen"))),
    ]
    for etiqueta, valor in datos:
        ws.cell(fila, 2, etiqueta).font = est["grupo_fuente"]
        c = ws.cell(fila, 3, valor)
        c.alignment = est["ajuste"]
        fila += 1

    fila += 1
    ws.cell(fila, 2, "Registro de Modificaciones").font = est["grupo_fuente"]
    fila += 1
    cabs = ["Versión", "Descripción", "Autor", "Fecha"]
    for i, cab in enumerate(cabs):
        c = ws.cell(fila, 2 + i, cab)
        c.font, c.fill, c.border, c.alignment = (
            est["cabecera_fuente"], est["cabecera_relleno"], est["borde"], est["centro"])
    registros = m.get("control_versiones") or [
        {"version": str(m.get("version", "1.0")), "descripcion": "Versión inicial",
         "autor": m.get("autor", ""), "fecha": hoy}]
    for r in registros:
        fila += 1
        for i, clave in enumerate(("version", "descripcion", "autor", "fecha")):
            c = ws.cell(fila, 2 + i, str(r.get(clave, "")))
            c.border, c.alignment, c.font = est["borde"], est["ajuste"], est["celda_fuente"]
    ws.freeze_panes = "A2"


def hoja_especificaciones(wb, m: dict, casos: list[dict], codigos: list[str],
                          est: dict) -> None:
    """El inventario. Es la hoja que importa: todo lo demas se deriva de aqui."""
    from openpyxl.utils import get_column_letter as L

    ws = wb.create_sheet("Especificaciones")
    ws.sheet_view.showGridLines = False

    ws.cell(4, 2, "NUM casos de prueba").font = est["grupo_fuente"]
    ws.cell(4, 3, len(casos)).font = est["grupo_fuente"]

    # Banda de grupos, como en la plantilla: donde acaba la identificacion y
    # empieza la especificacion. Orienta en una hoja de diecisiete columnas.
    for col, etiqueta in ((2, "Identificación y clasificación"),
                          (10, "Especificación detallada"),
                          (14, "Trazabilidad y carga masiva")):
        c = ws.cell(6, col, etiqueta)
        c.font, c.fill = est["grupo_fuente"], est["grupo_relleno"]

    for i, (nombre, ancho) in enumerate(COLUMNAS):
        col = COL_INI + i
        ws.column_dimensions[L(col)].width = ancho
        c = ws.cell(FILA_CAB, col, nombre)
        c.font, c.fill, c.border, c.alignment = (
            est["cabecera_fuente"], est["cabecera_relleno"], est["borde"], est["centro"])

    hu = f"{m.get('hu_id','')} - {m.get('titulo','')}".strip(" -")
    nombres_cu = {cu.get("id"): cu.get("nombre", "") for cu in m.get("casos_uso") or []}

    for n, (caso, codigo) in enumerate(zip(casos, codigos)):
        fila = FILA_DATOS + n
        cu = caso.get("caso_uso", "")
        valores = [
            NIVEL,
            hu,
            texto(f"{cu} - {nombres_cu.get(cu, '')}".strip(" -")),
            "",
            codigo,
            texto(caso.get("nombre")),
            texto(caso.get("criticidad")) or "Media",
            "Pendiente",
            texto(caso.get("descripcion")),
            texto(caso.get("precondiciones")),
            texto(caso.get("pasos")),
            texto(caso.get("resultado_esperado")),
            lista(caso.get("requisitos")),
            texto(caso.get("ejecucion")) or "manual",
            texto(caso.get("change")),
            texto(caso.get("duracion")),
            "",
        ]
        for i, v in enumerate(valores):
            c = ws.cell(fila, COL_INI + i, v)
            c.font, c.border, c.alignment = est["celda_fuente"], est["borde"], est["ajuste"]
        ws.cell(fila, COL_INI + 4).alignment = est["centro"]  # el codigo, centrado

    ultima = FILA_DATOS + max(len(casos), 1) - 1
    for valores, col in ((CRITICIDADES, 8), (ESTADOS, 9),
                         (EJECUCIONES, 15), (["S", "N"], 18)):
        _validacion(ws, valores, f"{L(col)}{FILA_DATOS}:{L(col)}{ultima + 50}")
    _validacion(ws, [NIVEL], f"B{FILA_DATOS}:B{ultima + 50}")

    ws.freeze_panes = f"B{FILA_DATOS}"
    ws.auto_filter.ref = f"B{FILA_CAB}:{L(COL_INI + len(COLUMNAS) - 1)}{ultima}"


def hoja_parametros(wb, est: dict) -> None:
    """La nomenclatura, visible. En la plantilla esta oculta y protegida.

    Aqui no hay macro que dependa de ella, asi que ocultarla solo escondaria de
    quien lee el plan la regla con la que se han formado los codigos.
    """
    ws = wb.create_sheet("Parámetros")
    ws.sheet_view.showGridLines = False
    for col, ancho in (("A", 3), ("B", 34), ("C", 10), ("D", 10), ("E", 46)):
        ws.column_dimensions[col].width = ancho

    ws.cell(2, 2, "PARAMETRIZACIÓN DE LA NOMENCLATURA").font = est["titulo_fuente"]
    ws.cell(4, 2, "Patrón").font = est["grupo_fuente"]
    ws.cell(4, 3, "AA.BB.CU.nn")
    ws.cell(5, 2, "Ejemplo").font = est["grupo_fuente"]
    ws.cell(5, 3, "PS.FU.CU01.01")

    filas = [
        ("AA", "Nivel de prueba", "PS = Pruebas de Sistema"),
        ("BB", "Subnivel", "FU = Funcionales"),
        ("CU", "Caso de uso", "El identificador del caso de uso que agrupa la prueba"),
        ("nn", "Correlativo", "Reinicia en 01 dentro de cada caso de uso"),
    ]
    ws.cell(7, 2, "Tramo").font = est["cabecera_fuente"]
    for i, cab in enumerate(("Tramo", "Significado", "Detalle")):
        c = ws.cell(7, 2 + i, cab)
        c.font, c.fill, c.border, c.alignment = (
            est["cabecera_fuente"], est["cabecera_relleno"], est["borde"], est["centro"])
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 60
    for n, (tramo, sig, det) in enumerate(filas):
        for i, v in enumerate((tramo, sig, det)):
            c = ws.cell(8 + n, 2 + i, v)
            c.border, c.alignment, c.font = est["borde"], est["ajuste"], est["celda_fuente"]

    ws.cell(14, 2, "Nivel de prueba cubierto por este plan").font = est["grupo_fuente"]
    ws.cell(15, 2, NIVEL)
    ws.cell(16, 2, "Los criterios de aceptación de una historia son pruebas funcionales "
                   "de sistema. Otros niveles (unitarias, integración, aceptación) no se "
                   "derivan de la historia y no se generan aquí.").alignment = est["ajuste"]


EJEC_COLS = [
    ("Código", 18), ("Nombre del caso de prueba", 40), ("Entorno", 12),
    ("Plataforma", 14), ("Sistema", 14), ("Navegador", 14),
    ("Versión navegador", 16), ("Fecha", 12), ("Resultado", 15),
    ("Ejecutado por", 20), ("Incidencia", 16), ("Observaciones", 40),
]
EJEC_CAB = 5
EJEC_INI = 6


def hoja_ejecucion(wb, m: dict, casos: list[dict], codigos: list[str], est: dict) -> str:
    """La rejilla donde se anota lo ejecutado. Sale con una fila por caso.

    En la plantilla esta hoja la creaba una macro pidiendo plan, plataforma y
    build; aqui se genera con los casos ya puestos y el entorno del manifiesto.
    """
    from openpyxl.utils import get_column_letter as L
    from openpyxl.worksheet.hyperlink import Hyperlink

    nombre = "Ejecución 1"
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False

    ws.cell(2, 2, "INVENTARIO DE EJECUCIONES DE PRUEBA").font = est["titulo_fuente"]
    ws.cell(3, 2, "Total pruebas").font = est["grupo_fuente"]
    ws.cell(3, 3, len(casos))
    ws.cell(3, 4, "Entorno").font = est["grupo_fuente"]
    ws.cell(3, 5, m.get("entorno", ""))

    for i, (cab, ancho) in enumerate(EJEC_COLS):
        col = COL_INI + i
        ws.column_dimensions[L(col)].width = ancho
        c = ws.cell(EJEC_CAB, col, cab)
        c.font, c.fill, c.border, c.alignment = (
            est["cabecera_fuente"], est["cabecera_relleno"], est["borde"], est["centro"])

    for n, (caso, codigo) in enumerate(zip(casos, codigos)):
        fila = EJEC_INI + n
        # El codigo enlaza a su fila en el inventario: en un plan de cien casos,
        # buscar a mano de que trata `PS.FU.CU07.03` es la friccion que sobra.
        c = ws.cell(fila, COL_INI, codigo)
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"Especificaciones!F{FILA_DATOS + n}")
        c.style = "Hyperlink"
        c.alignment = est["centro"]
        ws.cell(fila, COL_INI + 1, caso.get("nombre", "")).alignment = est["ajuste"]
        ws.cell(fila, COL_INI + 2, m.get("entorno", "")).alignment = est["centro"]
        ws.cell(fila, COL_INI + 8, "No ejecutado").alignment = est["centro"]
        for i in range(len(EJEC_COLS)):
            celda = ws.cell(fila, COL_INI + i)
            celda.border = est["borde"]
            if celda.font.size is None:
                celda.font = est["celda_fuente"]

    ultima = EJEC_INI + max(len(casos), 1) - 1
    for fila in range(EJEC_INI, ultima + 51):
        ws.cell(fila, COL_INI + 7).number_format = "DD/MM/YYYY"
    _validacion(ws, RESULTADOS, f"{L(COL_INI + 8)}{EJEC_INI}:{L(COL_INI + 8)}{ultima + 50}")
    ws.freeze_panes = f"B{EJEC_INI}"
    ws.auto_filter.ref = f"B{EJEC_CAB}:{L(COL_INI + len(EJEC_COLS) - 1)}{ultima}"
    return nombre


def hoja_qmetry(wb, m: dict, casos: list[dict], codigos: list[str], est: dict) -> None:
    """Proyeccion del inventario al formato de importacion de la herramienta.

    Es una vista, no una fuente: si el inventario cambia, se regenera el plan.
    """
    from openpyxl.utils import get_column_letter as L

    ws = wb.create_sheet("Qmetry")
    ws.sheet_view.showGridLines = False
    cabs = [("Summary", 46), ("Priority", 12), ("Status", 14), ("Description", 52),
            ("Step Summary", 52), ("Expected Result", 52), ("Labels", 18)]
    prioridad = {"Alta": "High", "Media": "Medium", "Baja": "Low"}

    for i, (cab, ancho) in enumerate(cabs):
        ws.column_dimensions[L(1 + i)].width = ancho
        c = ws.cell(1, 1 + i, cab)
        c.font, c.fill, c.border, c.alignment = (
            est["cabecera_fuente"], est["cabecera_relleno"], est["borde"], est["centro"])

    for n, (caso, codigo) in enumerate(zip(casos, codigos)):
        etiquetas = [e for e in (m.get("hu_id"), caso.get("ejecucion"), caso.get("change")) if e]
        valores = [
            f"{codigo} {caso.get('nombre', '')}".strip(),
            prioridad.get(caso.get("criticidad", "Media"), "Medium"),
            "To Do",
            texto(caso.get("descripcion")),
            texto(caso.get("pasos")),
            texto(caso.get("resultado_esperado")),
            " ".join(etiquetas),
        ]
        for i, v in enumerate(valores):
            c = ws.cell(2 + n, 1 + i, v)
            c.font, c.border, c.alignment = est["celda_fuente"], est["borde"], est["ajuste"]
    ws.freeze_panes = "A2"


def hoja_resumen(wb, casos: list[dict], est: dict, hoja_ejec: str) -> None:
    """KPIs con formulas vivas sobre la hoja de ejecucion.

    Con formula y no con el numero calculado aqui: quien ejecuta anota en la
    rejilla y el resumen se actualiza solo. Un recuento congelado a fecha de
    generacion siempre diria cero.
    """
    from openpyxl.utils import get_column_letter as L

    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False
    for col, ancho in (("A", 3), ("B", 26), ("C", 12), ("D", 12), ("E", 40)):
        ws.column_dimensions[col].width = ancho

    total = max(len(casos), 1)
    fin = EJEC_INI + total - 1 + 50  # el mismo margen que los desplegables
    col = L(COL_INI + 8)
    col_res = f"'{hoja_ejec}'!${col}${EJEC_INI}:${col}${fin}"

    ws.cell(2, 2, "RESUMEN DE EJECUCIÓN").font = est["titulo_fuente"]
    for i, cab in enumerate(("Resultado", "Casos", "%")):
        c = ws.cell(4, 2 + i, cab)
        c.font, c.fill, c.border, c.alignment = (
            est["cabecera_fuente"], est["cabecera_relleno"], est["borde"], est["centro"])

    for n, estado in enumerate(RESULTADOS):
        fila = 5 + n
        ws.cell(fila, 2, estado).font = est["celda_fuente"]
        ws.cell(fila, 3, f'=COUNTIF({col_res},"{estado}")')
        ws.cell(fila, 4, f"=IFERROR(C{fila}/$C$9,0)").number_format = "0%"
        for i in range(3):
            ws.cell(fila, 2 + i).border = est["borde"]
        ws.cell(fila, 3).alignment = est["centro"]
        ws.cell(fila, 4).alignment = est["centro"]

    ws.cell(9, 2, "Total").font = est["grupo_fuente"]
    ws.cell(9, 3, f'=COUNTA({col_res})').font = est["grupo_fuente"]
    for i in range(3):
        ws.cell(9, 2 + i).border = est["borde"]
    ws.cell(9, 3).alignment = est["centro"]

    manual = sum(1 for c in casos if c.get("ejecucion", "manual") != "automatizable")
    ws.cell(11, 2, "Reparto por ejecución").font = est["grupo_fuente"]
    for n, (etiqueta, valor) in enumerate((("Manual", manual),
                                           ("Automatizable", len(casos) - manual))):
        ws.cell(12 + n, 2, etiqueta).font = est["celda_fuente"]
        ws.cell(12 + n, 3, valor).alignment = est["centro"]
        for i in range(2):
            ws.cell(12 + n, 2 + i).border = est["borde"]
    ws.cell(15, 2, "El reparto dimensiona el trabajo manual de cada regresión. "
                   "No lo ejecuta nadie automáticamente: es información de planificación.")
    ws.cell(15, 2).alignment = est["ajuste"]


# --- Libro -------------------------------------------------------------------

def build(m: dict, salida: Path) -> dict:
    import branding as marca
    from openpyxl import Workbook

    m = sanear(m)
    br = marca.normalizar(m.get("branding"))
    est = marca.estilos_excel(br)
    hoy = m.get("fecha") or date.today().isoformat()

    casos = list(m.get("casos") or [])
    declarados = {cu.get("id") for cu in m.get("casos_uso") or []}
    huerfanos = sorted({c.get("caso_uso") for c in casos
                        if c.get("caso_uso") and c.get("caso_uso") not in declarados})

    _TRUNCADOS.clear()
    codigos = generar_codigos(casos)
    avisos = detectar_colisiones(casos, codigos)
    if huerfanos:
        avisos.append("casos de uso usados y no declarados en `casos_uso`: "
                      + ", ".join(huerfanos) + " (saldran sin nombre en el inventario)")

    wb = Workbook()
    hoja_control(wb, m, est, br, hoy, marca)
    hoja_especificaciones(wb, m, casos, codigos, est)
    hoja_parametros(wb, est)
    nombre_ejec = hoja_ejecucion(wb, m, casos, codigos, est)
    hoja_qmetry(wb, m, casos, codigos, est)
    hoja_resumen(wb, casos, est, nombre_ejec)

    for recorte in _TRUNCADOS:
        avisos.append(f"texto recortado al limite de una celda de Excel: '{recorte}'")

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    return {
        "salida": str(salida),
        "casos": len(casos),
        "casos_uso": len(declarados),
        "manual": sum(1 for c in casos if c.get("ejecucion", "manual") != "automatizable"),
        "automatizable": sum(1 for c in casos if c.get("ejecucion") == "automatizable"),
        "con_change": sum(1 for c in casos if c.get("change")),
        "marca": bool(br.get("activa")),
        "avisos": avisos,
    }


def comprobar(ruta: Path) -> dict:
    """Que hay ya anotado en un plan existente, antes de plantearse regenerarlo.

    La regla del skill --nunca pisar resultados ejecutados-- necesita un dato
    que solo esta dentro del `.xlsx`. Sin este modo, la regla se queda escrita
    y nadie puede cumplirla.
    """
    import openpyxl

    if not ruta.is_file():
        return {"fichero": str(ruta), "existe": False, "regenerable": True,
                "motivo": "no existe todavia"}

    wb = openpyxl.load_workbook(ruta, data_only=True)
    if "Ejecución 1" not in wb.sheetnames:
        return {"fichero": str(ruta), "existe": True, "regenerable": False,
                "motivo": "no tiene hoja 'Ejecución 1': no lo ha generado este skill, "
                          "asi que regenerarlo destruiria un fichero ajeno"}

    ws = wb["Ejecución 1"]
    col = COL_INI + 8
    recuento: dict[str, int] = {}
    total = 0
    for fila in range(EJEC_INI, ws.max_row + 1):
        if not ws.cell(fila, COL_INI).value:
            continue
        total += 1
        r = str(ws.cell(fila, col).value or "").strip()
        if r and r != "No ejecutado":
            recuento[r] = recuento.get(r, 0) + 1

    ejecutados = sum(recuento.values())
    return {
        "fichero": str(ruta),
        "existe": True,
        "casos": total,
        "ejecutados": ejecutados,
        "resultados": recuento,
        "regenerable": ejecutados == 0,
        "motivo": ("sin resultados anotados" if ejecutados == 0 else
                   f"{ejecutados} de {total} casos tienen resultado anotado; "
                   f"regenerar los borraria y no se recuperan"),
    }


def main() -> None:
    ap = argparse.ArgumentParser(description="Genera el Plan de Pruebas (.xlsx) de una HU.")
    ap.add_argument("--manifest", help="JSON con el contenido del plan; sin el, stdin")
    ap.add_argument("--output", help="ruta del .xlsx de salida")
    ap.add_argument("--schema", action="store_true",
                    help="imprime el esquema del manifiesto y sale")
    ap.add_argument("--comprobar", metavar="RUTA",
                    help="dice si un plan ya existente se puede regenerar sin "
                         "perder resultados anotados, y sale")
    ap.add_argument("--no-install", action="store_true",
                    help="no instalar openpyxl al vuelo")
    args = ap.parse_args()

    if args.schema:
        print(SCHEMA)
        return
    if args.comprobar:
        _ensure_openpyxl(not args.no_install)
        res = comprobar(Path(args.comprobar))
        print(json.dumps(res, ensure_ascii=False, indent=2))
        sys.exit(0 if res["regenerable"] else 1)
    if not args.output:
        ap.error("--output es obligatorio (o usa --schema)")

    _ensure_openpyxl(not args.no_install)

    crudo = (Path(args.manifest).read_text(encoding="utf-8") if args.manifest
             else sys.stdin.read())
    try:
        m = json.loads(crudo)
    except json.JSONDecodeError as e:
        sys.stderr.write(f"El manifiesto no es JSON valido: {e}\n")
        sys.exit(2)

    if not m.get("casos"):
        sys.stderr.write(
            "El manifiesto no trae ningun caso de prueba. Un plan vacio no es un "
            "entregable: revisa los criterios de aceptacion de la historia.\n")
        sys.exit(2)

    res = build(m, Path(args.output))
    for a in res["avisos"]:
        sys.stderr.write(f"Aviso: {a}\n")
    print(json.dumps(res, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
