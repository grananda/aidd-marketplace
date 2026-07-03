#!/usr/bin/env python3
"""Build the HU-review planning workbook (.xlsx) from a JSON build manifest.

This is the *rendering* half of the ``aidd-hu-review-plan`` skill. The skill (an
LLM) reads ``docs/mapa-historias-usuario.md`` and ``docs/detalle-historias-usuario.md``,
consolidates them, decides the July review schedule and writes a machine-readable
manifest (JSON). This script turns that manifest into an Excel workbook with four
sheets:

  * "Detalle HU"  — the consolidated user-story table. The "Historia
    (Como...quiero...para...)" column bolds the connector words Como/quiero/para
    via in-cell rich text.
  * "Dashboard"   — auto-computed KPIs (pending, closed, blocked...) plus native
    Excel charts (state, phase, priority, persona, review type).
  * "Leyenda"     — one block per field that needs explaining (Persona, GAP,
    Estado...), value -> meaning.
  * "Gantt Julio" — a day-by-day calendar of the review: week 1 = client-doc
    review, the rest of the month = per-HU definition/validation split into
    functional (business) and technical (TI) review meetings.

The Markdown produced by the skill remains the single source of truth; this
workbook is a complementary human-facing deliverable and never edits the source
docs.

Usage:
    python gen_hu_plan_xlsx.py --input docs/plan-revision-hu.json \
                               --output docs/xlsx/plan-revision-hu.xlsx [--open]
"""

from __future__ import annotations

import argparse
import calendar
import json
import subprocess
import sys
import webbrowser
from collections import Counter, OrderedDict
from datetime import date, timedelta
from pathlib import Path


def _ensure_openpyxl() -> None:
    """Import openpyxl, installing it on the fly if missing.

    This skill is meant to be run by non-technical users, so a missing
    dependency must not stop them. If the import fails we try to install
    openpyxl with pip (first normally, then with --user for restricted
    environments) and retry. Opt out with the ``--no-install`` flag or the
    ``AIDD_HU_PLAN_NO_INSTALL`` env var. The outer ``python`` invocation is the
    permission gate; the pip subprocess inherits that approval.
    """
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass

    import os
    if "--no-install" in sys.argv or os.environ.get("AIDD_HU_PLAN_NO_INSTALL"):
        sys.stderr.write(
            "ERROR: falta el paquete 'openpyxl' y la autoinstalacion esta desactivada.\n"
            "Instalalo con:  pip install openpyxl\n"
        )
        sys.exit(2)

    sys.stderr.write("Aviso: 'openpyxl' no esta instalado; instalandolo automaticamente...\n")
    attempts = (
        [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
        [sys.executable, "-m", "pip", "install", "--quiet", "--user", "openpyxl"],
    )
    for cmd in attempts:
        try:
            subprocess.check_call(cmd)
        except Exception:  # noqa: BLE001 - try the next strategy
            continue
        try:
            import openpyxl  # noqa: F401
            sys.stderr.write("OK: 'openpyxl' instalado correctamente.\n")
            return
        except ImportError:
            continue

    sys.stderr.write(
        "ERROR: no se pudo instalar 'openpyxl' automaticamente.\n"
        "Instalalo manualmente con:  pip install openpyxl   (o  python -m pip install openpyxl)\n"
    )
    sys.exit(2)


_ensure_openpyxl()

from openpyxl import Workbook  # noqa: E402
from openpyxl.cell.rich_text import CellRichText, TextBlock  # noqa: E402
from openpyxl.cell.text import InlineFont  # noqa: E402
from openpyxl.chart import BarChart, DoughnutChart, Reference  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


# --------------------------------------------------------------------------- #
# Palette / styling helpers
# --------------------------------------------------------------------------- #

BRAND = "1A3D7C"          # NTT DATA-ish deep blue
BRAND_LIGHT = "DCE5F3"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=BRAND, size=16)
SUB_FONT = Font(italic=True, color="666666", size=9)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATE_COLORS = {
    "Cerrada": "63BE7B",
    "En revision": "FFD966",
    "En revisión": "FFD966",
    "Pendiente": "9DC3E6",
    "Bloqueada": "E67C73",
}
PRIORITY_COLORS = {"Alta": "E67C73", "Media": "FFD966", "Baja": "9DC3E6"}
GANTT_COLORS = {
    "doc": "9DC3E6",        # client-doc review week
    "funcional": "A9D08E",  # business review
    "tecnica": "F4B183",    # TI review
    "técnica": "F4B183",
    "otro": "D9D9D9",
}
GANTT_LETTER = {"funcional": "F", "tecnica": "T", "técnica": "T"}
WEEKEND_FILL = PatternFill("solid", fgColor="EFEFEF")
KPI_FILL = PatternFill("solid", fgColor=BRAND_LIGHT)


def norm_state(hu: dict) -> str:
    """Derive a normalized state, honouring the explicit blocked flag."""
    if hu.get("bloqueada"):
        return "Bloqueada"
    st = (hu.get("estado") or "Pendiente").strip()
    if st.lower().startswith("cerrad"):
        return "Cerrada"
    if st.lower().startswith("en revis"):
        return "En revision"
    if st.lower().startswith("bloque"):
        return "Bloqueada"
    return st or "Pendiente"


def style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


# --------------------------------------------------------------------------- #
# Sheet: Detalle HU
# --------------------------------------------------------------------------- #

DETALLE_COLS = [
    ("id", "ID", 10),
    ("fase", "Fase", 7),
    ("persona", "Persona", 9),
    ("epica", "Epica / Actividad", 20),
    ("historia", "Historia (Como...quiero...para...)", 55),
    ("prioridad", "Prioridad", 10),
    ("moscow", "MoSCoW", 9),
    ("estimacion", "Estim.", 7),
    ("estado", "Estado", 12),
    ("gap", "GAP", 10),
    ("rf", "RF", 9),
    ("tipo_revision", "Revision", 11),
    ("dependencias", "Dependencias", 16),
    ("criterios", "Criterios de aceptacion", 50),
    ("notas", "Notas tecnicas", 32),
]


def historia_richtext(hu: dict) -> CellRichText:
    """Compose 'Como <..> quiero <..> para <..>' with connectors in bold."""
    b = InlineFont(b=True)
    parts: list = []
    como = (hu.get("como") or "").strip()
    quiero = (hu.get("quiero") or "").strip()
    para = (hu.get("para") or "").strip()
    if como:
        parts += [TextBlock(b, "Como "), como]
    if quiero:
        parts += [TextBlock(b, " quiero "), quiero]
    if para:
        parts += [TextBlock(b, " para "), para]
    if not parts:
        # Fallback: a pre-composed sentence; bold the keywords in place.
        raw = (hu.get("historia") or "").strip()
        return CellRichText([raw]) if raw else CellRichText([""])
    return CellRichText(parts)


def build_detalle(ws, hus: list) -> None:
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Detalle de Historias de Usuario").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Consolidado de mapa-historias-usuario.md + detalle-historias-usuario.md. "
                  "Fuente de verdad: docs/plan-revision-hu.md").font = SUB_FONT

    hdr = 4
    for i, (_, label, width) in enumerate(DETALLE_COLS, start=1):
        ws.cell(row=hdr, column=i, value=label)
        ws.column_dimensions[get_column_letter(i)].width = width
    style_header_row(ws, hdr, len(DETALLE_COLS))

    for r, hu in enumerate(hus, start=hdr + 1):
        state = norm_state(hu)
        for i, (key, _, _) in enumerate(DETALLE_COLS, start=1):
            cell = ws.cell(row=r, column=i)
            cell.border = BORDER
            cell.alignment = LEFT
            if key == "historia":
                cell.value = historia_richtext(hu)
            elif key == "estado":
                cell.value = state
                color = STATE_COLORS.get(state)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = CENTER
            elif key == "prioridad":
                cell.value = hu.get("prioridad", "")
                color = PRIORITY_COLORS.get((hu.get("prioridad") or "").strip())
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = CENTER
            elif key in ("dependencias", "criterios"):
                val = hu.get(key) or []
                if isinstance(val, list):
                    cell.value = "\n".join(f"- {x}" for x in val) if val else ""
                else:
                    cell.value = str(val)
            elif key in ("fase", "persona", "moscow", "estimacion", "gap", "rf",
                         "tipo_revision"):
                cell.value = hu.get(key, "")
                cell.alignment = CENTER
            else:
                cell.value = hu.get(key, "")

    ws.freeze_panes = ws.cell(row=hdr + 1, column=2)
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(DETALLE_COLS))}{hdr + len(hus)}"


# --------------------------------------------------------------------------- #
# Sheet: Dashboard
# --------------------------------------------------------------------------- #

def _write_kpi(ws, row, col, label, value):
    lc = ws.cell(row=row, column=col, value=value)
    lc.font = Font(bold=True, size=22, color=BRAND)
    lc.alignment = CENTER
    lc.fill = KPI_FILL
    tc = ws.cell(row=row + 1, column=col, value=label)
    tc.font = Font(size=9, color="444444")
    tc.alignment = CENTER
    tc.fill = KPI_FILL
    for rr in (row, row + 1):
        ws.cell(row=rr, column=col).border = BORDER


def _write_count_table(ws, top, col, title, counter: "OrderedDict[str, int]"):
    """Write a 2-col category/value table; return (data_ref, cats_ref, next_row)."""
    ws.cell(row=top, column=col, value=title).font = BOLD
    ws.cell(row=top, column=col + 1, value="Nº").font = BOLD
    r = top + 1
    for name, n in counter.items():
        ws.cell(row=r, column=col, value=name)
        ws.cell(row=r, column=col + 1, value=n)
        r += 1
    cats = Reference(ws, min_col=col, min_row=top + 1, max_row=r - 1)
    data = Reference(ws, min_col=col + 1, min_row=top, max_row=r - 1)
    return data, cats, r + 1


def _ordered_counter(values, preferred=None) -> "OrderedDict[str, int]":
    c = Counter(v for v in values if str(v).strip())
    out: "OrderedDict[str, int]" = OrderedDict()
    if preferred:
        for k in preferred:
            if k in c:
                out[k] = c.pop(k)
    for k in sorted(c):
        out[k] = c[k]
    return out


def build_dashboard(ws, hus: list) -> None:
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Cuadro de mando — Revision de HU").font = TITLE_FONT
    ws.cell(row=2, column=1, value="KPIs y graficas auto-calculadas del detalle de HU.").font = SUB_FONT

    total = len(hus)
    states = [norm_state(h) for h in hus]
    cerradas = states.count("Cerrada")
    bloqueadas = states.count("Bloqueada")
    en_rev = states.count("En revision")
    pendientes = total - cerradas
    pct = f"{round(100 * cerradas / total)}%" if total else "0%"

    kpis = [
        ("Total HU", total),
        ("Pendientes de cerrar", pendientes),
        ("Cerradas", cerradas),
        ("En revision", en_rev),
        ("Bloqueadas", bloqueadas),
        ("% Cerradas", pct),
    ]
    for i, (label, value) in enumerate(kpis):
        col = 1 + i * 2
        _write_kpi(ws, 4, col, label, value)
        ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions[get_column_letter(col + 1)].width = 6

    # Helper tables (kept to the right, columns R+) feeding the charts.
    helper_col = 20
    top = 4
    d1, c1, top = _write_count_table(ws, top, helper_col, "Estado",
                                     _ordered_counter(states,
                                                      ["Pendiente", "En revision", "Cerrada", "Bloqueada"]))
    d2, c2, top = _write_count_table(ws, top, helper_col, "Fase",
                                     _ordered_counter([h.get("fase") for h in hus],
                                                      ["F0", "F1", "F2", "F3"]))
    d3, c3, top = _write_count_table(ws, top, helper_col, "Prioridad",
                                     _ordered_counter([h.get("prioridad") for h in hus],
                                                      ["Alta", "Media", "Baja"]))
    d4, c4, top = _write_count_table(ws, top, helper_col, "Persona",
                                     _ordered_counter([h.get("persona") for h in hus]))
    d5, c5, top = _write_count_table(ws, top, helper_col, "Tipo de revision",
                                     _ordered_counter([h.get("tipo_revision") for h in hus],
                                                      ["funcional", "tecnica", "ambas"]))
    ws.column_dimensions[get_column_letter(helper_col)].width = 16

    def _doughnut(title, data, cats, anchor):
        ch = DoughnutChart()
        ch.title = title
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ch.height = 7
        ch.width = 10
        ws.add_chart(ch, anchor)

    def _bar(title, data, cats, anchor):
        ch = BarChart()
        ch.type = "col"
        ch.title = title
        ch.legend = None
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ch.height = 7
        ch.width = 10
        ws.add_chart(ch, anchor)

    _doughnut("HU por estado", d1, c1, "A8")
    _bar("HU por fase", d2, c2, "H8")
    _bar("HU por prioridad", d3, c3, "A24")
    _bar("HU por persona", d4, c4, "H24")
    _bar("HU por tipo de revision", d5, c5, "A40")


# --------------------------------------------------------------------------- #
# Sheet: Leyenda
# --------------------------------------------------------------------------- #

def build_leyenda(ws, legend: list) -> None:
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Leyenda de campos").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Significado de los valores codificados. Completar los que queden en blanco.").font = SUB_FONT
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    r = 4
    if not legend:
        ws.cell(row=r, column=1, value="(sin campos codificados detectados)").font = SUB_FONT
        return
    for block in legend:
        campo = block.get("campo", "")
        ws.cell(row=r, column=1, value=f"Campo: {campo}").font = Font(bold=True, size=12, color=BRAND)
        r += 1
        ws.cell(row=r, column=1, value="Valor")
        ws.cell(row=r, column=2, value="Significado")
        style_header_row(ws, r, 2)
        r += 1
        for item in block.get("valores", []):
            ws.cell(row=r, column=1, value=item.get("valor", "")).border = BORDER
            mc = ws.cell(row=r, column=2, value=item.get("significado", ""))
            mc.border = BORDER
            mc.alignment = LEFT
            ws.cell(row=r, column=1).alignment = CENTER
            r += 1
        r += 1  # blank line between blocks


# --------------------------------------------------------------------------- #
# Sheet: Gantt Julio
# --------------------------------------------------------------------------- #

WEEKDAY_ABBR = ["L", "M", "X", "J", "V", "S", "D"]  # Mon..Sun


def build_gantt(ws, gantt: dict) -> None:
    ws.sheet_view.showGridLines = False
    year = int(gantt.get("year"))
    month = int(gantt.get("month"))
    ndays = calendar.monthrange(year, month)[1]
    month_name = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                  "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"][month]

    ws.cell(row=1, column=1, value=f"Gantt — Planificacion de revision de HU ({month_name} {year})").font = TITLE_FONT
    kickoff = gantt.get("kickoff")
    if kickoff:
        ws.cell(row=2, column=1, value=f"Kickoff: {kickoff}. Semana 1: revision de documentacion del cliente. "
                                       "Resto del mes: definicion y validacion de HU (reuniones funcionales/TI).").font = SUB_FONT

    label_cols = 2                # A:B for the activity label
    day0 = label_cols + 1         # first day column
    hdr = 4                       # day-number header row
    wk = 5                        # weekday-abbreviation row

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    for d in range(1, ndays + 1):
        col = day0 + d - 1
        ws.column_dimensions[get_column_letter(col)].width = 3.6
        dt = date(year, month, d)
        weekend = dt.weekday() >= 5
        n = ws.cell(row=hdr, column=col, value=d)
        w = ws.cell(row=wk, column=col, value=WEEKDAY_ABBR[dt.weekday()])
        for cell in (n, w):
            cell.alignment = CENTER
            cell.border = BORDER
            cell.font = Font(bold=True, size=8, color="FFFFFF" if not weekend else "888888")
            cell.fill = HEADER_FILL if not weekend else WEEKEND_FILL
    ws.cell(row=hdr, column=1, value="Actividad").font = HEADER_FONT
    ws.cell(row=hdr, column=1).fill = HEADER_FILL
    ws.cell(row=hdr, column=1).alignment = CENTER
    ws.merge_cells(start_row=hdr, start_column=1, end_row=wk, end_column=2)

    row = wk + 1
    for entry in gantt.get("rows", []):
        kind = (entry.get("kind") or "otro").strip().lower()
        color = GANTT_COLORS.get(kind, GANTT_COLORS["otro"])
        letter = GANTT_LETTER.get(kind, "")
        label = ws.cell(row=row, column=1, value=entry.get("activity", ""))
        label.alignment = LEFT
        label.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=2).border = BORDER

        start = entry.get("start")
        end = entry.get("end") or start
        meetings = set(entry.get("meetings") or [])
        try:
            sd = date.fromisoformat(start) if start else None
            ed = date.fromisoformat(end) if end else sd
        except ValueError:
            sd = ed = None

        # weekend shading across the whole row first
        for d in range(1, ndays + 1):
            col = day0 + d - 1
            dt = date(year, month, d)
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            if dt.weekday() >= 5:
                cell.fill = WEEKEND_FILL

        if sd and ed:
            cur = sd
            while cur <= ed:
                if cur.month == month and cur.year == year and cur.weekday() < 5:
                    col = day0 + cur.day - 1
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.border = BORDER
                    if cur.isoformat() in meetings and letter:
                        cell.value = letter
                        cell.font = Font(bold=True, size=8, color="333333")
                        cell.alignment = CENTER
                cur += timedelta(days=1)
        ws.row_dimensions[row].height = 16
        row += 1

    # Gantt legend
    row += 1
    ws.cell(row=row, column=1, value="Leyenda:").font = BOLD
    legend_items = [
        ("doc", "Revision documentacion cliente"),
        ("funcional", "Reunion funcional (negocio) — F"),
        ("tecnica", "Reunion tecnica (TI) — T"),
    ]
    for i, (kind, text) in enumerate(legend_items):
        c = ws.cell(row=row + 1 + i, column=1)
        c.fill = PatternFill("solid", fgColor=GANTT_COLORS[kind])
        c.border = BORDER
        ws.cell(row=row + 1 + i, column=2, value=text).alignment = LEFT
        ws.merge_cells(start_row=row + 1 + i, start_column=2, end_row=row + 1 + i, end_column=6)

    ws.freeze_panes = ws.cell(row=wk + 1, column=day0)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def load_manifest(path: Path) -> dict:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("El manifiesto JSON debe ser un objeto en la raiz.")
    if not isinstance(data.get("hus"), list) or not data["hus"]:
        raise ValueError("El manifiesto debe incluir una lista no vacia 'hus'.")
    data.setdefault("legend", [])
    data.setdefault("gantt", {})
    return data


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description="Genera el Excel de planificacion de revision de HU.")
    p.add_argument("--input", required=True, help="Manifiesto JSON de entrada.")
    p.add_argument("--output", required=True, help="Ruta del .xlsx de salida.")
    p.add_argument("--open", action="store_true", help="Abrir el .xlsx al terminar (best-effort).")
    p.add_argument("--no-install", action="store_true",
                   help="No autoinstalar openpyxl si falta (por defecto se instala).")
    args = p.parse_args(argv)

    in_path = Path(args.input)
    out_path = Path(args.output)
    if not in_path.is_file():
        sys.stderr.write(f"ERROR: no existe el manifiesto de entrada: {in_path}\n")
        return 2
    try:
        manifest = load_manifest(in_path)
    except (ValueError, json.JSONDecodeError) as exc:
        sys.stderr.write(f"ERROR: manifiesto invalido: {exc}\n")
        return 2

    hus = manifest["hus"]
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    build_detalle(wb.create_sheet("Detalle HU"), hus)
    build_dashboard(wb.create_sheet("Dashboard"), hus)
    build_leyenda(wb.create_sheet("Leyenda"), manifest["legend"])
    gantt = manifest.get("gantt") or {}
    if gantt.get("year") and gantt.get("month"):
        build_gantt(wb.create_sheet("Gantt Julio"), gantt)

    # Put Dashboard first — it is the landing view.
    wb.move_sheet("Dashboard", -wb.sheetnames.index("Dashboard"))
    wb.active = 0

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f"OK  -> {out_path}")
    print(f"    HU: {len(hus)}  |  hojas: {', '.join(wb.sheetnames)}")
    if args.open:
        try:
            webbrowser.open(out_path.resolve().as_uri())
        except Exception:  # noqa: BLE001 - never fail the render on open
            pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
